"""
Stock Prediction Automation System - Railway Version
Runs once daily at 9:30 AM EST for 2 weeks
"""

import os
import sys
import yfinance as yf
import pandas as pd
from openai import OpenAI
from datetime import datetime, timedelta
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
import time
import json
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# Load environment variables from .env file (for local development)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv not installed, will use system environment variables

# Configuration
TICKERS = ["TSLA", "NVDA", "AMZN", "META", "AAPL"]
OUTPUT_FILE = "predictions.xlsx"
LOG_FILE = "run_log.txt"
CONFIG_FILE = "config.json"
MODEL = "gpt-4o"  # Change to "gpt-5" or "gpt-5o" if GPT-5 becomes available. Use --check-model flag to verify.

# Get API key from environment variable
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY", "")

# Email configuration
EMAIL_ADDRESS = "aadityasai.kalyankar@gmail.com"
EMAIL_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")  # Gmail App Password from Railway

client = None


def get_openai_client():
    global client
    if client is None:
        if not OPENAI_API_KEY:
            raise ValueError("Set OPENAI_API_KEY environment variable in Railway")
        client = OpenAI(api_key=OPENAI_API_KEY)
    return client


def log(message: str):
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    log_msg = f"[{timestamp}] {message}"
    print(log_msg)
    with open(LOG_FILE, "a") as f:
        f.write(log_msg + "\n")


def send_email(subject: str, body: str, attachment_path: str = None):
    """Send email with optional Excel attachment"""
    if not EMAIL_PASSWORD:
        log("GMAIL_APP_PASSWORD not set. Skipping email.")
        return
    
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = EMAIL_ADDRESS
        msg['To'] = EMAIL_ADDRESS
        msg['Subject'] = subject
        
        # Add body
        msg.attach(MIMEText(body, 'plain'))
        
        # Add attachment if provided
        if attachment_path and os.path.exists(attachment_path):
            with open(attachment_path, "rb") as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header(
                    'Content-Disposition',
                    f'attachment; filename= {os.path.basename(attachment_path)}'
                )
                msg.attach(part)
        
        # Send email
        server = smtplib.SMTP('smtp.gmail.com', 587)
        server.starttls()
        server.login(EMAIL_ADDRESS, EMAIL_PASSWORD)
        text = msg.as_string()
        server.sendmail(EMAIL_ADDRESS, EMAIL_ADDRESS, text)
        server.quit()
        
        log(f"Email sent successfully: {subject}")
    except Exception as e:
        log(f"Error sending email: {e}")


def get_config():
    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r") as f:
            return json.load(f)
    else:
        config = {
            "start_date": "2026-01-14",
            "end_date": "2026-02-18",
            "final_report_date": "2026-02-19",
            "run_count": 0,
            "max_runs": 25,  # Trading days: Jan 14-17, 21-24, 27-31, Feb 3-7, 10-14, 17-18 (skips Jan 20 MLK Day, Feb 16 Presidents Day)
            "final_report_generated": False
        }
        save_config(config)
        return config


def save_config(config):
    with open(CONFIG_FILE, "w") as f:
        json.dump(config, f, indent=2)


def is_experiment_complete(config) -> bool:
    today = datetime.now().strftime("%Y-%m-%d")
    if today > config["end_date"]:
        return True
    return config["run_count"] >= config["max_runs"]


def is_market_day() -> bool:
    today = datetime.now()
    if today.weekday() > 4:
        return False
    
    holidays = [
        "2025-01-01", "2025-01-20", "2025-02-17", "2025-04-18",
        "2025-05-26", "2025-06-19", "2025-07-04", "2025-09-01",
        "2025-11-27", "2025-12-25",
        "2026-01-01", "2026-01-19", "2026-01-20", "2026-02-16", "2026-04-03",
        "2026-05-25", "2026-06-19", "2026-07-03", "2026-09-07",
        "2026-11-26", "2026-12-25",
    ]
    
    return today.strftime("%Y-%m-%d") not in holidays


def get_stock_prices(ticker: str) -> dict:
    try:
        stock = yf.Ticker(ticker)
        # Get recent data for today's prices
        hist_recent = stock.history(period="5d")
        
        if len(hist_recent) < 2:
            log(f"Warning: Not enough data for {ticker}")
            return None
        
        # Get extended historical data (6 months) for trend analysis
        hist_extended = stock.history(period="6mo")
        
        today_open = hist_recent['Open'].iloc[-1]
        today_close = hist_recent['Close'].iloc[-1]
        yesterday_close = hist_recent['Close'].iloc[-2]
        closing_prices_recent = hist_recent['Close'].tail(10).tolist()
        
        # Get historical closing prices over 6 months (sampled weekly for manageable size)
        if len(hist_extended) > 0:
            # Sample approximately weekly data points (every ~5 trading days)
            historical_prices = hist_extended['Close'].iloc[::5].tolist()
            # Ensure we have at least 20 data points, if not, use all available
            if len(historical_prices) < 20:
                historical_prices = hist_extended['Close'].tolist()
        else:
            historical_prices = closing_prices_recent
        
        return {
            "ticker": ticker,
            "today_open": round(today_open, 2),
            "today_close": round(today_close, 2),
            "yesterday_close": round(yesterday_close, 2),
            "closing_prices": [round(p, 2) for p in closing_prices_recent],
            "historical_prices": [round(p, 2) for p in historical_prices]
        }
    except Exception as e:
        log(f"Error fetching {ticker}: {e}")
        return None


def create_prompts(ticker: str, closing_prices: list, historical_prices: list) -> dict:
    price_list = ", ".join([f"${p}" for p in closing_prices])
    historical_price_list = ", ".join([f"${p}" for p in historical_prices])
    
    return {
        "prompt_1": {
            "name": "Basic",
            "text": f"""For the stock ticker {ticker}, predict the direction of the stock price movement for the next trading day.

You MUST choose either UP or DOWN - neutral is not an option.
Respond with ONLY one of the following options:
UP
DOWN
Do not include explanations, numbers, probabilities, or additional text."""
        },
        
        "prompt_2": {
            "name": "Price Data",
            "text": f"""You are given EXTENDED HISTORICAL closing prices (over the past 6 months) for the stock ticker {ticker}.
These historical closing prices span multiple months and show long-term price trends (most recent last):
{historical_price_list}

Analyze the EXTENDED historical price trend pattern over time. Look at the overall trend direction, momentum, and price behavior patterns across the entire historical period. Based ONLY on this extended historical numerical price pattern and trend analysis, predict the direction of the stock's movement for the NEXT trading day (tomorrow).
Consider the longer-term trend direction, not just recent short-term fluctuations.

You MUST choose either UP or DOWN - neutral is not an option.
Respond with ONLY one of the following options:
UP
DOWN
Do not include explanations, indicators, probabilities, or any additional text."""
        },
        
        "prompt_3": {
            "name": "Research",
            "text": f"""For the stock ticker {ticker}, conduct COMPREHENSIVE DETAILED RESEARCH and analysis:

1. HISTORICAL PRICE DATA:
Recent closing prices from previous trading days (most recent last):
{price_list}

Extended historical closing prices over the past 6 months (most recent last):
{historical_price_list}

2. COMPREHENSIVE RESEARCH - Analyze ALL of the following:
- Recent financial news headlines, earnings reports, and earnings history
- Analyst ratings, price targets, and recommendations
- Company fundamentals (revenue trends, profitability, growth metrics)
- Industry trends and sector performance
- Market conditions and broader economic factors
- Technical indicators and price patterns
- Recent corporate actions (mergers, acquisitions, partnerships, product launches)
- Regulatory changes or policy impacts
- Competitive positioning and market share
- Management commentary and guidance
- Any significant events or catalysts from the past week

Using BOTH the comprehensive historical price trend analysis AND your detailed research across all these factors, synthesize a well-informed prediction for the direction of the stock's movement for the next trading day.

You MUST choose either UP or DOWN based on this comprehensive analysis - neutral is not an option.
Respond with ONLY one of the following options:
UP
DOWN
Do not include explanations, sources, probabilities, numbers, or any additional text."""
        }
    }


def list_available_models():
    """List all models available to your API account"""
    try:
        log("=" * 60)
        log("CHECKING AVAILABLE MODELS")
        log("=" * 60)
        
        # Get list of available models
        models = get_openai_client().models.list()
        
        log("Available models in your account:")
        gpt_models = []
        for model in models.data:
            if "gpt" in model.id.lower():
                gpt_models.append(model.id)
                log(f"  - {model.id}")
        
        if not gpt_models:
            log("  No GPT models found in your account")
        else:
            log(f"\nTotal GPT models available: {len(gpt_models)}")
            
            # Check for GPT-5
            gpt5_models = [m for m in gpt_models if "gpt-5" in m.lower() or "gpt5" in m.lower()]
            if gpt5_models:
                log(f"\n✓ GPT-5 MODELS FOUND: {', '.join(gpt5_models)}")
            else:
                log("\n⚠ No GPT-5 models found. Latest available: " + max(gpt_models, key=lambda x: x.lower()))
        
        log("=" * 60)
        return gpt_models
        
    except Exception as e:
        log(f"Error listing models: {e}")
        log("Note: Some API keys may not have permission to list models")
        return None


def check_model_version():
    """Check which model is actually being used by the API"""
    try:
        log("=" * 60)
        log("CHECKING MODEL VERSION")
        log("=" * 60)
        log(f"Configured model: {MODEL}")
        
        # Make a test API call to see what model is actually used
        response = get_openai_client().chat.completions.create(
            model=MODEL,
            messages=[
                {"role": "user", "content": "Say 'test'"}
            ],
            max_tokens=5
        )
        
        # Get the actual model used from the response
        actual_model = response.model
        log(f"Actual model used by API: {actual_model}")
        
        # Check if it's GPT-5
        is_gpt5 = "gpt-5" in actual_model.lower() or "gpt5" in actual_model.lower()
        
        if is_gpt5:
            log("✓ CONFIRMED: Using GPT-5!")
        elif "gpt-4" in actual_model.lower():
            log("⚠ Using GPT-4 (not GPT-5)")
        else:
            log(f"ℹ Using model: {actual_model}")
        
        log("=" * 60)
        return actual_model
        
    except Exception as e:
        log(f"Error checking model version: {e}")
        if "does not exist" in str(e) or "not found" in str(e).lower():
            log(f"\n⚠ Model '{MODEL}' is not available to your account.")
            log("Run with --list-models to see available models.")
        return None


def get_prediction(prompt: str) -> str:
    try:
        response = get_openai_client().chat.completions.create(
            model=MODEL,
            messages=[
                {"role": "system", "content": "You are a financial analyst. You must respond with ONLY UP or DOWN. Neutral is not an option."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=10,
            temperature=0.3
        )
        
        # Log the actual model used (for verification)
        actual_model = response.model
        if actual_model != MODEL:
            log(f"Note: Requested {MODEL}, but API used {actual_model}")
        
        result = response.choices[0].message.content.strip().upper()
        
        # Only allow UP or DOWN - no NEUTRAL
        for valid in ["UP", "DOWN"]:
            if valid in result:
                return valid
        # If neither found, default to UP (since neutral is not allowed)
        log(f"Warning: Invalid response '{result}', defaulting to UP")
        return "UP"
        
    except Exception as e:
        log(f"API Error: {e}")
        return "ERROR"


def initialize_excel():
    if os.path.exists(OUTPUT_FILE):
        return
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"
    
    headers = ["Day #", "Date", "Ticker", "Open", "Close", "Prompt 1", "Prompt 2", "Prompt 3", "Actual", "P1 ✓", "P2 ✓", "P3 ✓"]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    wb.save(OUTPUT_FILE)
    log(f"Created {OUTPUT_FILE}")


def run_daily():
    config = get_config()
    
    # Check if it's Feb 19th and generate final report
    today = datetime.now().strftime("%Y-%m-%d")
    if today == config.get("final_report_date", "2026-02-19") and not config.get("final_report_generated", False):
        if os.path.exists(OUTPUT_FILE):
            generate_final_excel_report()
            config['final_report_generated'] = True
            save_config(config)
            log("Final report generated. Experiment complete!")
        return
    
    if is_experiment_complete(config):
        log("=" * 60)
        log("EXPERIMENT COMPLETE! 2 weeks of data collected.")
        log("=" * 60)
        return
    
    if not is_market_day():
        log("Market closed today. Skipping.")
        return
    
    log("=" * 60)
    log(f"DAY {config['run_count'] + 1} of {config['max_runs']}")
    log("=" * 60)
    
    initialize_excel()
    
    wb = load_workbook(OUTPUT_FILE)
    ws = wb.active
    
    today = datetime.now().strftime("%Y-%m-%d")
    day_num = config['run_count'] + 1
    
    for ticker in TICKERS:
        log(f"\n{ticker}:")
        
        prices = get_stock_prices(ticker)
        if not prices:
            continue
        
        log(f"  Today Open: ${prices['today_open']}")
        log(f"  Today Close: ${prices['today_close']}")
        log(f"  Yesterday Close: ${prices['yesterday_close']}")
        
        if prices['today_open'] > prices['yesterday_close']:
            actual = "UP"
        elif prices['today_open'] < prices['yesterday_close']:
            actual = "DOWN"
        else:
            # If equal, default to UP (since neutral is not an option)
            actual = "UP"
        
        log(f"  Actual: {actual}")
        
        prompts = create_prompts(ticker, prices['closing_prices'], prices.get('historical_prices', prices['closing_prices']))
        
        predictions = {}
        for key, prompt_data in prompts.items():
            pred = get_prediction(prompt_data['text'])
            predictions[key] = pred
            log(f"  {prompt_data['name']}: {pred}")
            time.sleep(0.5)
        
        p1_correct = "✓" if predictions['prompt_1'] == actual else "✗"
        p2_correct = "✓" if predictions['prompt_2'] == actual else "✗"
        p3_correct = "✓" if predictions['prompt_3'] == actual else "✗"
        
        next_row = ws.max_row + 1
        row_data = [
            day_num, today, ticker,
            prices['today_open'], prices['today_close'],
            predictions['prompt_1'], predictions['prompt_2'], predictions['prompt_3'],
            actual, p1_correct, p2_correct, p3_correct
        ]
        
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=next_row, column=col, value=value)
            cell.alignment = Alignment(horizontal='center')
            if col >= 10:
                if value == "✓":
                    cell.font = Font(color="008000", bold=True)
                else:
                    cell.font = Font(color="FF0000")
    
    wb.save(OUTPUT_FILE)
    
    config['run_count'] += 1
    save_config(config)
    
    log(f"\nDay {day_num} complete! Runs until Feb 18th.")
    
    # Send daily email with results
    email_subject = f"Stock Prediction Results - Day {day_num} ({today})"
    email_body = f"""Stock Prediction Experiment - Daily Update

Day: {day_num} of {config['max_runs']}
Date: {today}

Predictions completed for all {len(TICKERS)} tickers:
{', '.join(TICKERS)}

The Excel file with today's predictions is attached.

Experiment runs until Feb 18th, 2026.
"""
    send_email(email_subject, email_body, OUTPUT_FILE)


def generate_final_excel_report():
    """Generate final Excel report with summary statistics on Feb 19th"""
    if not os.path.exists(OUTPUT_FILE):
        log("No data file found. Cannot generate final report.")
        return
    
    df = pd.read_excel(OUTPUT_FILE)
    
    if df.empty:
        log("No data collected yet. Cannot generate final report.")
        return
    
    log("=" * 60)
    log("GENERATING FINAL EXCEL REPORT")
    log("=" * 60)
    
    # Load existing workbook
    wb = load_workbook(OUTPUT_FILE)
    
    # Create or clear Summary sheet
    if "Summary" in wb.sheetnames:
        wb.remove(wb["Summary"])
    ws_summary = wb.create_sheet("Summary", 0)  # Insert at beginning
    
    # Summary header
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    title_font = Font(bold=True, size=16)
    
    ws_summary.merge_cells('A1:D1')
    title_cell = ws_summary.cell(row=1, column=1, value="STOCK PREDICTION EXPERIMENT - FINAL RESULTS")
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    
    # Experiment details
    total = len(df)
    days = df['Day #'].nunique()
    start_date = df['Date'].min()
    end_date = df['Date'].max()
    
    ws_summary.cell(row=3, column=1, value="Experiment Period:").font = Font(bold=True)
    ws_summary.cell(row=3, column=2, value=f"{start_date} to {end_date}")
    ws_summary.cell(row=4, column=1, value="Total Predictions:").font = Font(bold=True)
    ws_summary.cell(row=4, column=2, value=total)
    ws_summary.cell(row=5, column=1, value="Trading Days:").font = Font(bold=True)
    ws_summary.cell(row=5, column=2, value=days)
    
    # Results table header
    row = 7
    headers = ["Prompt Type", "Correct", "Total", "Accuracy %"]
    for col, header in enumerate(headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Calculate accuracies
    row += 1
    prompt_names = [
        ('P1 ✓', 'Prompt 1 (Basic)'),
        ('P2 ✓', 'Prompt 2 (Historical Price Data)'),
        ('P3 ✓', 'Prompt 3 (Comprehensive Research)')
    ]
    
    for col_name, display_name in prompt_names:
        correct = (df[col_name] == '✓').sum()
        accuracy = (correct / total) * 100
        
        ws_summary.cell(row=row, column=1, value=display_name).font = Font(bold=True)
        ws_summary.cell(row=row, column=2, value=correct).alignment = Alignment(horizontal='center')
        ws_summary.cell(row=row, column=3, value=total).alignment = Alignment(horizontal='center')
        accuracy_cell = ws_summary.cell(row=row, column=4, value=f"{accuracy:.2f}%")
        accuracy_cell.alignment = Alignment(horizontal='center')
        accuracy_cell.font = Font(bold=True)
        row += 1
    
    # Per-ticker breakdown
    row += 2
    ws_summary.cell(row=row, column=1, value="PER-TICKER ACCURACY").font = Font(bold=True, size=12)
    row += 1
    
    ticker_headers = ["Ticker", "P1 Accuracy", "P2 Accuracy", "P3 Accuracy"]
    for col, header in enumerate(ticker_headers, 1):
        cell = ws_summary.cell(row=row, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    row += 1
    for ticker in TICKERS:
        ticker_df = df[df['Ticker'] == ticker]
        if len(ticker_df) > 0:
            ws_summary.cell(row=row, column=1, value=ticker).font = Font(bold=True)
            for i, col_name in enumerate(['P1 ✓', 'P2 ✓', 'P3 ✓'], 2):
                correct = (ticker_df[col_name] == '✓').sum()
                accuracy = (correct / len(ticker_df)) * 100
                acc_cell = ws_summary.cell(row=row, column=i, value=f"{accuracy:.2f}%")
                acc_cell.alignment = Alignment(horizontal='center')
            row += 1
    
    # Auto-adjust column widths
    for col_idx in range(1, 5):  # Columns A-D
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for row_idx in range(1, ws_summary.max_row + 1):
            cell = ws_summary.cell(row=row_idx, column=col_idx)
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, 50)
        ws_summary.column_dimensions[column_letter].width = adjusted_width
    
    # Save final report
    final_report_file = "final_report_feb18.xlsx"
    wb.save(final_report_file)
    log(f"Final Excel report saved: {final_report_file}")
    
    # Also print summary
    print("\n" + "=" * 60)
    print("FINAL RESULTS")
    print("=" * 60)
    print(f"\nPredictions: {total} | Days: {days}")
    
    summary_text = f"\nPredictions: {total} | Days: {days}\n\n"
    for col_name, display_name in prompt_names:
        correct = (df[col_name] == '✓').sum()
        accuracy = (correct / total) * 100
        result_line = f"{display_name}: {correct}/{total} = {accuracy:.1f}%"
        print(result_line)
        summary_text += result_line + "\n"
    
    # Send final report email
    email_subject = "Stock Prediction Experiment - FINAL REPORT (Feb 18th)"
    email_body = f"""Stock Prediction Experiment - FINAL RESULTS

Experiment Period: {start_date} to {end_date}
Total Predictions: {total}
Trading Days: {days}

ACCURACY RESULTS:
{summary_text}

The complete final Excel report with summary statistics and per-ticker breakdown is attached.

Experiment Complete!
"""
    send_email(email_subject, email_body, final_report_file)


def generate_report():
    """Legacy function - prints report to console"""
    if not os.path.exists(OUTPUT_FILE):
        print("No data file found.")
        return
    
    df = pd.read_excel(OUTPUT_FILE)
    
    if df.empty:
        print("No data collected yet.")
        return
    
    print("\n" + "=" * 60)
    print("FINAL RESULTS")
    print("=" * 60)
    
    total = len(df)
    days = df['Day #'].nunique()
    
    print(f"\nPredictions: {total} | Days: {days}")
    
    for col, name in [('P1 ✓', 'Prompt 1 (Basic)'), ('P2 ✓', 'Prompt 2 (Historical Price Data)'), ('P3 ✓', 'Prompt 3 (Comprehensive Research)')]:
        correct = (df[col] == '✓').sum()
        accuracy = (correct / total) * 100
        print(f"{name}: {correct}/{total} = {accuracy:.1f}%")


if __name__ == "__main__":
    if len(sys.argv) > 1 and sys.argv[1] == "--report":
        generate_report()
    elif len(sys.argv) > 1 and sys.argv[1] == "--check-model":
        check_model_version()
    elif len(sys.argv) > 1 and sys.argv[1] == "--list-models":
        list_available_models()
    else:
        # Check model version on first run
        if not os.path.exists(CONFIG_FILE):
            check_model_version()
        run_daily()

"""
Test script to generate a sample final Excel report
"""
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# Create sample predictions data
TICKERS = ["TSLA", "NVDA", "AMZN", "META", "AAPL"]
sample_data = []

# Generate 13 days of sample data (Jan 14-17, 21-24, 27-31)
dates = [
    "2026-01-14", "2026-01-15", "2026-01-16", "2026-01-17",
    "2026-01-21", "2026-01-22", "2026-01-23", "2026-01-24",
    "2026-01-27", "2026-01-28", "2026-01-29", "2026-01-30", "2026-01-31"
]

day_num = 1
for date in dates:
    for ticker in TICKERS:
        # Random sample predictions and results
        import random
        predictions = random.choices(["UP", "DOWN"], k=3)
        actual = random.choice(["UP", "DOWN"])
        
        p1_correct = "✓" if predictions[0] == actual else "✗"
        p2_correct = "✓" if predictions[1] == actual else "✗"
        p3_correct = "✓" if predictions[2] == actual else "✗"
        
        sample_data.append({
            "Day #": day_num,
            "Date": date,
            "Ticker": ticker,
            "Open": round(random.uniform(100, 500), 2),
            "Close": round(random.uniform(100, 500), 2),
            "Prompt 1": predictions[0],
            "Prompt 2": predictions[1],
            "Prompt 3": predictions[2],
            "Actual": actual,
            "P1 ✓": p1_correct,
            "P2 ✓": p2_correct,
            "P3 ✓": p3_correct
        })
    day_num += 1

# Create initial Excel file
df = pd.DataFrame(sample_data)
df.to_excel("predictions.xlsx", index=False, sheet_name="Predictions")

# Now generate the final report (similar to main.py)
wb = load_workbook("predictions.xlsx")

# Create Summary sheet
if "Summary" in wb.sheetnames:
    wb.remove(wb["Summary"])
ws_summary = wb.create_sheet("Summary", 0)

# Styling
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=14)
title_font = Font(bold=True, size=16)

# Title
ws_summary.merge_cells('A1:D1')
title_cell = ws_summary.cell(row=1, column=1, value="STOCK PREDICTION EXPERIMENT - FINAL RESULTS")
title_cell.font = title_font
title_cell.alignment = Alignment(horizontal='center')

# Experiment details
total = len(df)
days = df['Day #'].nunique()
start_date = df['Date'].min()
end_date = df['Date'].max()

ws_summary.cell(row=3, column=1, value="Experiment Period:").font = Font(bold=True)
ws_summary.cell(row=3, column=2, value=f"{start_date} to {end_date}")
ws_summary.cell(row=4, column=1, value="Total Predictions:").font = Font(bold=True)
ws_summary.cell(row=4, column=2, value=total)
ws_summary.cell(row=5, column=1, value="Trading Days:").font = Font(bold=True)
ws_summary.cell(row=5, column=2, value=days)

# Results table header
row = 7
headers = ["Prompt Type", "Correct", "Total", "Accuracy %"]
for col, header in enumerate(headers, 1):
    cell = ws_summary.cell(row=row, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

# Calculate accuracies
row += 1
prompt_names = [
    ('P1 ✓', 'Prompt 1 (Basic)'),
    ('P2 ✓', 'Prompt 2 (Price Data)'),
    ('P3 ✓', 'Prompt 3 (Price Data + News)')
]

for col_name, display_name in prompt_names:
    correct = (df[col_name] == '✓').sum()
    accuracy = (correct / total) * 100
    
    ws_summary.cell(row=row, column=1, value=display_name).font = Font(bold=True)
    ws_summary.cell(row=row, column=2, value=correct).alignment = Alignment(horizontal='center')
    ws_summary.cell(row=row, column=3, value=total).alignment = Alignment(horizontal='center')
    accuracy_cell = ws_summary.cell(row=row, column=4, value=f"{accuracy:.2f}%")
    accuracy_cell.alignment = Alignment(horizontal='center')
    accuracy_cell.font = Font(bold=True)
    row += 1

# Per-ticker breakdown
row += 2
ws_summary.cell(row=row, column=1, value="PER-TICKER ACCURACY").font = Font(bold=True, size=12)
row += 1

ticker_headers = ["Ticker", "P1 Accuracy", "P2 Accuracy", "P3 Accuracy"]
for col, header in enumerate(ticker_headers, 1):
    cell = ws_summary.cell(row=row, column=col, value=header)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

row += 1
for ticker in TICKERS:
    ticker_df = df[df['Ticker'] == ticker]
    if len(ticker_df) > 0:
        ws_summary.cell(row=row, column=1, value=ticker).font = Font(bold=True)
        for i, col_name in enumerate(['P1 ✓', 'P2 ✓', 'P3 ✓'], 2):
            correct = (ticker_df[col_name] == '✓').sum()
            accuracy = (correct / len(ticker_df)) * 100
            acc_cell = ws_summary.cell(row=row, column=i, value=f"{accuracy:.2f}%")
            acc_cell.alignment = Alignment(horizontal='center')
        row += 1

# Auto-adjust column widths
for col_idx in range(1, 5):  # Columns A-D
    max_length = 0
    column_letter = get_column_letter(col_idx)
    for row_idx in range(1, ws_summary.max_row + 1):
        cell = ws_summary.cell(row=row_idx, column=col_idx)
        if cell.value:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
    adjusted_width = min(max_length + 2, 50)
    ws_summary.column_dimensions[column_letter].width = adjusted_width

# Format the Predictions sheet
ws_predictions = wb["Predictions"]
header_fill_pred = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_pred = Font(bold=True, color="FFFFFF")

# Format header row
for col in range(1, len(df.columns) + 1):
    cell = ws_predictions.cell(row=1, column=col)
    cell.fill = header_fill_pred
    cell.font = header_font_pred
    cell.alignment = Alignment(horizontal='center')

# Format data rows
for row_idx in range(2, len(df) + 2):
    for col_idx in range(1, len(df.columns) + 1):
        cell = ws_predictions.cell(row=row_idx, column=col_idx)
        cell.alignment = Alignment(horizontal='center')
        # Color code accuracy columns (columns 10-12)
        if col_idx >= 10:
            if cell.value == "✓":
                cell.font = Font(color="008000", bold=True)
            elif cell.value == "✗":
                cell.font = Font(color="FF0000")

# Auto-adjust column widths for Predictions sheet
for col_idx in range(1, len(df.columns) + 1):
    max_length = 0
    column_letter = get_column_letter(col_idx)
    for row_idx in range(1, ws_predictions.max_row + 1):
        cell = ws_predictions.cell(row=row_idx, column=col_idx)
        if cell.value:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
    adjusted_width = min(max_length + 2, 30)
    ws_predictions.column_dimensions[column_letter].width = adjusted_width

# Save final report
final_report_file = "final_report_mar4.xlsx"
wb.save(final_report_file)
print(f"✅ Test Excel report created: {final_report_file}")
print(f"   - Summary sheet with overall and per-ticker statistics")
print(f"   - Predictions sheet with all daily data")

